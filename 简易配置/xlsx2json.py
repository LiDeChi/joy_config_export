import json
import os
import tkinter as tk
from tkinter import filedialog, Listbox, END
from openpyxl import load_workbook

def preprocess_data(value):
    if value == "null":
        return None
    if isinstance(value, str) and value.startswith("[") and value.endswith("]"):
        # If the value is a string representation of a list, convert it to a list
        return json.loads(value.replace("'", '"'))
    return value

def xlsx_to_json(file_path):
    try:
        # Load the workbook
        workbook = load_workbook(file_path)

        # Create a dictionary to store data from all worksheets
        data = {}

        # Iterate over each worksheet
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            # Skip sheets whose names start with "%"
            if sheet_name.startswith("%"):
                continue
            sheet = workbook[sheet_name]

            # If it's the first worksheet, handle it differently
            if sheet_name.lower() == "global":
                # Create a dictionary to store data from the first worksheet
                sheet_data = {}

                # Iterate over each row (starting from the second row)
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # Get the first column as the field name
                    field_name = row[0]

                    # Get the value from the second column
                    value = row[1]

                    # Skip the third column
                    if field_name == value == None:
                        continue

                    # If not the first row (header row) and the second row
                    if row_index >= 2:
                        # If value is not None, preprocess the value
                        if value is not None:
                            value = preprocess_data(value)
                        sheet_data[field_name] = value

                # Store the data from the first worksheet in the result
                data[sheet_name] = sheet_data
            else:
                # Get headers
                headers = [cell.value for cell in sheet[2]]

                # Create a list to store data from the current worksheet
                sheet_data = []

                # Iterate over each row (skipping the header row)
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    # Convert row data to a dictionary
                    row_data = dict(zip(headers, map(preprocess_data, row)))
                    sheet_data.append(row_data)

                # Store data from the current worksheet in the result
                data[sheet_name] = sheet_data

        # Remove any "null" keys that may have been retained during processing
        data = {k: v for k, v in data.items() if v is not None}

        # Convert the result to a JSON string
        json_data = json.dumps(data, ensure_ascii=False, separators=(',', ':'))

        # 获取输入文件的目录和文件名（不包括扩展名）
        file_dir = os.path.dirname(file_path)
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # 构建输出文件路径
        output_path = os.path.join(file_dir, f"{file_name}.json")

        # 将JSON数据写入文件
        with open(output_path, 'w', encoding='utf-8') as file:
            file.write(json_data)

        print(f"转换完成！JSON文件已保存到：{output_path}")
        return f"转换完成！JSON文件已保存到：{output_path}"
    except Exception as e:
        print(f"发生错误：{e}")
        return f"发生错误：{e}"

def load_last_path():
    if os.path.exists('last_path.json'):
        with open('last_path.json', 'r') as file:
            return json.load(file).get('last_path', '')
    return ''

def save_last_path(path):
    with open('last_path.json', 'w') as file:
        json.dump({'last_path': path}, file)

def select_files():
    last_path = load_last_path()
    file_paths = filedialog.askopenfilenames(initialdir=last_path, title="Select Excel Files",
                                             filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if file_paths:
        save_last_path(os.path.dirname(file_paths[0]))
        for file_path in file_paths:
            if file_path not in selected_files:
                selected_files.append(file_path)
                listbox.insert(END, file_path)

def convert_files():
    for idx, file_path in enumerate(selected_files):
        result = xlsx_to_json(file_path)
        listbox.insert(END, result)
    selected_files.clear()

selected_files = []

# Create the main window
root = tk.Tk()
root.title("Excel to JSON Converter")

# Create and place the listbox
listbox = Listbox(root, selectmode=tk.MULTIPLE, width=100, height=20)
listbox.pack(pady=20)

# Create and place the buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_button = tk.Button(button_frame, text="Select Files", command=select_files)
select_button.grid(row=0, column=0, padx=10)

convert_button = tk.Button(button_frame, text="Convert Files", command=convert_files)
convert_button.grid(row=0, column=1, padx=10)

# Run the main loop
root.mainloop()
